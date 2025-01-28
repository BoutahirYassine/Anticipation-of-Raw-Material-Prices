# -*- coding: utf-8 -*-
"""
Created on Tue Jan 28 14:22:04 2025

@author: boutahirya
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jan 16 13:39:43 2025

@author: boutahirya
"""

from bs4 import BeautifulSoup
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# URL of the page
url = "https://www.metal.com/fr/prices/202210280001"  # Replace with the actual URL

# Send a request to fetch the HTML content of the page
response = requests.get(url)

if response.status_code == 200:
    # Parse the HTML content with BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')
    # Locate the div containing "Données brutes"
    brut_section = soup.find('div', text='Données brutes')
    if brut_section:
        # Find the parent div and extract the price
        price_element = brut_section.find_next('div', class_='PriceDisplay_price__VYiMd')

    if price_element:
        # Extract the price
        price = price_element.text.strip()
        print(f"N-Type Polysilicon Price: {price} CNY/kg")

        # Get today's date
        today_date = datetime.now().strftime("%Y-%m-%d")

        # Define the Excel file name
        excel_file = "DataBI/Cell_TOPCON_M10.xlsx"

        # Check if the file exists
        if os.path.exists(excel_file):
            # Load the workbook if it exists
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        else:
            # Create a new workbook if it doesn't exist
            workbook = Workbook()
            sheet = workbook.active
            # Add headers to the file
            sheet.append(["Date", "Price (CNY/kg)"])

        # Check if the date already exists in the sheet
        date_found = False
        for idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, values_only=True), start=2):  # Skip header row
            if row[0] == today_date:  # Check if the date exists
                # Update the price if the date exists
                sheet.cell(row=idx, column=2, value=price)  # Update the price
                date_found = True
                print(f"Updated the price for date: {today_date}")

        # If the date was not found, append the new data
        if not date_found:
            sheet.append([today_date, price])
            print(f"Added new data for date: {today_date}")

        # Save the workbook
        workbook.save(excel_file)
        print(f"Data saved to {excel_file}")
    else:
        print("Price element not found.")
else:
    print(f"Failed to fetch the webpage. Status code: {response.status_code}")
