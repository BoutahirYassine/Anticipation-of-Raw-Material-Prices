from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import json
import csv
import openpyxl
import os
from datetime import datetime, timedelta

# URL cible
url1 = "https://fr.investing.com/currencies/usd-cny-historical-data"
file_name1 = 'DataBI/USD_CNY.xlsx'

url2 = "https://fr.investing.com/currencies/eur-cny-historical-data"
file_name2 = 'DataBI/EUR_CNY.xlsx'

url3 = "https://fr.investing.com/currencies/eur-usd-historical-data"
file_name3 = 'DataBI/EUR_USD.xlsx'

urlr = [[url1,file_name1],[url2,file_name2],[url3,file_name3]]
for tablee in urlr:
    url = tablee[0]
    file_name = tablee[1]
    # Append or update data in the Excel file
    def update_or_replace_data(sheet, data_to_add):
        date_to_update, rate_to_update = data_to_add
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from the second row
            existing_date = row[0]  # Value in column A
            if existing_date.value.date() == date_to_update.date():
                # Update the value in column B for the corresponding row
                sheet.cell(row=row_idx, column=2, value=rate_to_update)  # Column B = 2
                print(f"Updated rate for {date_to_update} in row {rate_to_update}")
                return
        # If the date doesn't exist, append it
        sheet.append(data_to_add)
        print(f"Added new rate for {date_to_update}: {rate_to_update}")
    
    
    # Configurer WebDriver
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Exécuter sans interface graphique
    service = Service("chromedriver-win64/chromedriver.exe")  # Remplacez par le chemin de chromedriver
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    # Charger la page
    driver.get(url)
    
    # Localiser la balise <script> contenant le JSON
    script_element = driver.find_element("xpath", "//script[@id='__NEXT_DATA__']")
    
    # Extraire le contenu brut du JSON
    raw_json = script_element.get_attribute("innerHTML")
    
    # Charger le JSON brut en tant qu'objet Python pour validation
    parsed_json = json.loads(raw_json)
    data = parsed_json
    
    # Navigate to the historical data
    historical_data = data['props']['pageProps']['state']['historicalDataStore']['historicalData']['data']
    
    today1 = datetime.now().strftime("%d/%m/%Y")
    yesterday1 = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")
    today = datetime.now()
    yesterday = (datetime.now() - timedelta(days=1))
    
    # Initialize values for last_close
    today_last_close = None
    yesterday_last_close = None
    
    # Extract last_close for today and yesterday
    for entry in historical_data:
        if entry.get('rowDate') == today1:
            today_last_close = entry.get('last_close')
        elif entry.get('rowDate') == yesterday1:
            yesterday_last_close = entry.get('last_close')
    
    data_to_add_today = [today, today_last_close]
    data_to_add_yesterday = [yesterday, yesterday_last_close]
    # Step 3: Update the Excel file
    
    
    # Check if the Excel file already exists
    if os.path.exists(file_name):
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active  # Use the active sheet
    else:
        # Create a new workbook and add headers
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Date', 'Rate'])  # Add headers
    
    # Append the new data (Date in column A, Rate in column B)
    update_or_replace_data(sheet, data_to_add_yesterday)
    update_or_replace_data(sheet, data_to_add_today)
    
    # Save the workbook
    workbook.save(file_name)
    
    
    driver.quit()
