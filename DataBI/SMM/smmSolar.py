# -*- coding: utf-8 -*-
"""
Created on Mon Mar  3 10:09:13 2025

@author: boutahirya
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jan 16 13:39:43 2025

@author: boutahirya
"""

from bs4 import BeautifulSoup
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

import requests
from bs4 import BeautifulSoup
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

def fetch_price(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        if url == "https://www.metal.com/Solar/202303220001":
            brut_section = soup.find('div', text='Original')
            if brut_section:              
                price_element = brut_section.find_next('div', class_='price___2mpJr')               # price WAFER
                
                
        elif url == "https://www.metal.com/en/prices/202210280001":
            brut_section = soup.find('div', text='Original')
            if brut_section:
                price_element = brut_section.find_next('div', class_='PriceDisplay_price__VYiMd')  #price CELL
                
                
        elif url == "https://www.metal.com/Solar/202112230003":
            brut_section = soup.find('div', text='Original')
            if brut_section:
                price_element = brut_section.find_next('div', class_='price___2mpJr')               #Silver Rear side
                
                
        elif url == "https://www.metal.com/Solar/202112230004":
            brut_section = soup.find('div', text='Original')
            if brut_section:
                price_element = brut_section.find_next('div', class_='price___2mpJr')              # Silver busbar front-side
                
                
        elif url == "https://www.metal.com/Solar/202112230005":
            brut_section = soup.find('div', text='Original')
            if brut_section:
                price_element = brut_section.find_next('div', class_='price___2mpJr')              #Silver Finger Front-side 
                
                
        # USD/CNY
        elif url == "https://www.metal.com/exchange-rate/200002250101":
            price_element1 = soup.find("span", class_="strong___Js3_I priceDown___2TbRQ")
            price_element2 = soup.find("span", class_="strong___Js3_I priceUp___3Mgsl")
            price_element3 = soup.find("span", class_="strong___3sC58 priceUp___3Mgsl")
            
            if price_element1:
                price_element = price_element1
            elif price_element3:
                price_element = price_element3
            else:
                price_element = price_element2
            
        # Find the parent div and extract the price MODULE
        elif url == "https://www.metal.com/Solar/202403260002":
            price_element = soup.find("span", class_="strong___3sC58")
        else:
            brut_section = soup.find('div', text='Original')
            if brut_section:
                price_element = brut_section.find_next('div', class_='price___2mpJr') 
                
        if price_element:
            return price_element.text.strip()
    return None

def update_excel(component_name, url, workbook):
    price = fetch_price(url)
    if price is None:
        print(f"Failed to fetch price for {component_name}")
        return

    today_date = datetime.now().strftime("%Y-%m-%d")
    sheet_name = "Prices"

    # Création de la feuille si elle n'existe pas
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Date", component_name])

    # Récupère les en-têtes (ligne 1)
    headers = [cell.value for cell in sheet[1]]
    if component_name not in headers:
        headers.append(component_name)
        sheet.cell(row=1, column=len(headers), value=component_name)

    # Trouver l'index de la colonne du composant
    col_idx = headers.index(component_name) + 1  # Excel = 1-based index

    # Vérifie si la date existe déjà dans une ligne
    date_found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == today_date:
            row[col_idx - 1].value = price
            date_found = True
            print(f"Updated {component_name} price on {today_date}")
            break

    if not date_found:
        new_row = [None] * len(headers)
        new_row[0] = today_date
        new_row[col_idx - 1] = price
        sheet.append(new_row)
        print(f"Added new row with {component_name} price on {today_date}")

def main():
    excel_file = "DataBI/SMM/Solar_Prices.xlsx"
    #excel_file = "Solar_Prices2.xlsx"
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
    
    update_excel("N-Dense", "https://www.metal.com/Solar/202501060001", workbook)
    update_excel("N-Type", "https://www.metal.com/Solar/202501060003", workbook)
    update_excel("N-type Wafer", "https://www.metal.com/Solar/202303220001", workbook) ##
    update_excel("Cell Topcon 183mm", "https://www.metal.com/en/prices/202210280001", workbook)
    update_excel("Module Topcon 183mm", "https://www.metal.com/Solar/202403260002", workbook)
    update_excel("Silver Rear_side", "https://www.metal.com/Solar/202112230003", workbook)
    update_excel("Silver Busbar front-side", "https://www.metal.com/Solar/202112230004", workbook)
    update_excel("Silver finger front-side", "https://www.metal.com/Solar/202112230005", workbook)
    update_excel("USD_CNY", "https://www.metal.com/exchange-rate/200002250101", workbook)
    
    workbook.save(excel_file)
    print(f"Data saved to {excel_file}")

if __name__ == "__main__":
    main()