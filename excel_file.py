import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
# Chemin du dossier contenant les fichiers Excel
folder_path = r'files'

# Chemin du fichier Excel de destination
output_file = r'DataBI/prix_matieres_output.xlsx'

# Créer des dictionnaires pour chaque source
infolink_data = {
    'Date': [],
    'Mono N Type Wafer - 182-183.75mm / 130µm (RMB)': [],
    'TOPCon Cell - 182-183.75mm / 24.9%+ (USD)': [],
    'TOPCon Cell - 182-183.75mm / 24.9%+ (RMB)': [],
    '182*182-210mm/210mm Mono TOPCon - EU (USD)' : []
}
energytrend_data = {
    'Date': [],
    'NType Polysilicon (RMB)': [],
    'NType M10 Mono Wafer -182mm/130μm (RMB)': [],
    'M10 TOPCon Cell (RMB)': [],
    '182mm TOPCon Module (RMB)': []
}
pvinsights_data = {
    'Date': [],
    '182.2mm x 183.75mm N Mono Wafer': [],
    '182mm N-Mono Cell': [],
    '182mm 580/625Wp N-Mono Module': [],
    'USD/CNY': []
    
}

# Parcourir tous les fichiers du dossier
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and (filename.startswith('infolink') or filename.startswith('energytrend') or filename.startswith('pvinsights')):
        file_path = os.path.join(folder_path, filename)
        
        # Extraire la date en fonction du préfixe
        if filename.startswith('infolink'):
            date_str = filename[9:].split('.')[0]
        
        elif filename.startswith('energytrend'):
            date_str = filename[12:].split('.')[0]
        
        elif filename.startswith('pvinsights'):
            date_str = filename[11:].split('.')[0]
        
        # Charger le fichier Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Extraire la valeur de la cellule en fonction du fichier et ajouter au dictionnaire approprié
        if filename.startswith('infolink'):
            infolink_value1 = sheet.cell(row=9, column=2).value
            infolink_value2 = sheet.cell(row=16, column=2).value
            infolink_value3 = sheet.cell(row=17, column=2).value
            infolink_value4 = sheet.cell(row=35, column=2).value
            infolink_data['Date'].append(date_str)
            infolink_data['Mono N Type Wafer - 182-183.75mm / 130µm (RMB)'].append(infolink_value1)
            infolink_data['TOPCon Cell - 182-183.75mm / 24.9%+ (USD)'].append(infolink_value2)
            infolink_data['TOPCon Cell - 182-183.75mm / 24.9%+ (RMB)'].append(infolink_value3)
            infolink_data['182*182-210mm/210mm Mono TOPCon - EU (USD)'].append(infolink_value4)
        
        elif filename.startswith('energytrend'):
            # Extraire la date du nom du fichier
            # Supposons que le format est "energytrend_YYYY-MM-DD.xlsx"
            file_date_str = filename.split('_')[1].split('.')[0]  # "2024-10-21"
            file_date = datetime.strptime(file_date_str, "%Y-%m-%d")  # Conversion en objet datetime
            
            # Date limite
            limit_date = datetime.strptime("2024-12-17", "%Y-%m-%d")
            
            # Vérifier si la date du fichier est avant la date limite
            if file_date <= limit_date:
                # Charger le fichier Excel
                workbook = load_workbook(filename)
                sheet = workbook.active  # Assurez-vous que la feuille active est correcte
                
                # Lire les valeurs des cellules spécifiques
                energytrend_value1 = sheet.cell(row=6, column=2).value
                energytrend_value2 = sheet.cell(row=13, column=2).value
                energytrend_value3 = sheet.cell(row=21, column=2).value
                energytrend_value4 = sheet.cell(row=29, column=2).value
            else : 
                # Charger le fichier Excel
                workbook = load_workbook(filename)
                sheet = workbook.active  # Assurez-vous que la feuille active est correcte
                
                # Lire les valeurs des cellules spécifiques
                energytrend_value1 = sheet.cell(row=4, column=2).value
                energytrend_value2 = sheet.cell(row=12, column=2).value
                energytrend_value3 = sheet.cell(row=20, column=2).value
                energytrend_value4 = sheet.cell(row=28, column=2).value

            energytrend_data['Date'].append(date_str)
            energytrend_data['NType Polysilicon (RMB)'].append(energytrend_value1)
            energytrend_data['NType M10 Mono Wafer -182mm/130μm (RMB)'].append(energytrend_value2)
            energytrend_data['M10 TOPCon Cell (RMB)'].append(energytrend_value3)
            energytrend_data['182mm TOPCon Module (RMB)'].append(energytrend_value4)
            
        elif filename.startswith('pvinsights'):
            pvinsights_value1 = sheet.cell(row=6, column=2).value
            pvinsights_value2 = sheet.cell(row=13, column=2).value
            pvinsights_value3 = sheet.cell(row=26, column=2).value
            pvinsights_value4 = sheet.cell(row=2, column=8).value
            pvinsights_data['Date'].append(date_str)
            pvinsights_data['182.2mm x 183.75mm N Mono Wafer'].append(pvinsights_value1)
            pvinsights_data['182mm N-Mono Cell'].append(pvinsights_value2)
            pvinsights_data['182mm 580/625Wp N-Mono Module'].append(pvinsights_value3)
            pvinsights_data['USD/CNY'].append(pvinsights_value4)

# Créer des DataFrames pour chaque source
df_infolink = pd.DataFrame(infolink_data)
df_energytrend = pd.DataFrame(energytrend_data)
df_pvinsights = pd.DataFrame(pvinsights_data)

# Écrire dans le fichier Excel avec chaque source dans une feuille distincte et ajuster la largeur des colonnes
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Enregistrement des DataFrames dans des feuilles séparées
    df_infolink.to_excel(writer, sheet_name='infolink', index=False)
    df_energytrend.to_excel(writer, sheet_name='energytrend', index=False)
    df_pvinsights.to_excel(writer, sheet_name='pvinsights', index=False)
    
    # Ajuster la largeur des colonnes pour chaque feuille
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Ajustement de la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Trouver la longueur maximale du contenu de chaque colonne
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajouter un peu d'espace supplémentaire
            worksheet.column_dimensions[column_letter].width = adjusted_width

