from bs4 import BeautifulSoup
import requests
import openpyxl
import os

# URL de la page contenant les cours de l'argent
url = "https://www.abcbourse.com/download/valeur/XAGUSDp"

# Envoyer une requête pour récupérer le contenu HTML
html = requests.get(url)
soup = BeautifulSoup(html.content, "html.parser")

# Trouver toutes les lignes du tableau
rows = soup.select("#tabQuotes tbody tr")

# Vérification si au moins deux valeurs existent
if len(rows) < 2:
    print("Erreur : Pas assez de données disponibles.")
    exit()

# Extraire les deux dernières valeurs
latest_row = rows[0]  # Dernière valeur
previous_row = rows[1]  # Avant-dernière valeur

# Extraire les données
latest_date = latest_row.find_all("td")[0].text.strip()
latest_price = latest_row.find_all("td")[5].text.strip()

previous_date = previous_row.find_all("td")[0].text.strip()
previous_price = previous_row.find_all("td")[5].text.strip()

# Afficher les résultats
print(f"Dernière valeur ({latest_date}): {latest_price}")
print(f"Avant-dernière valeur ({previous_date}): {previous_price}")

# Nom du fichier Excel
file_name = "DataBI/Argent.xlsx"

# Vérifier si le fichier existe, sinon le créer
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Silver Prices"
    ws.append(["Date", "Prix"])
    wb.save(file_name)

# Charger le fichier Excel existant
wb = openpyxl.load_workbook(file_name)
ws = wb.active

# Fonction pour ajouter ou mettre à jour les données
def update_or_add_entry(date, price):
    updated = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[0].value == date:  # Vérifier si la date existe déjà
            row[1].value = price  # Mettre à jour la valeur existante
            updated = True
            break
    if not updated:  # Si la date n'existe pas, ajouter une nouvelle ligne
        ws.append([date, price])

# Ajouter ou mettre à jour les valeurs des deux dernières dates
update_or_add_entry(latest_date, latest_price)
update_or_add_entry(previous_date, previous_price)

# Sauvegarder les modifications dans le fichier Excel
wb.save(file_name)
print("Données mises à jour dans", file_name)
