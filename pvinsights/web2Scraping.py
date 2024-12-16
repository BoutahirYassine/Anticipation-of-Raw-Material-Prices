import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import certifi

# Requête vers la page web
url = 'https://pvinsights.com/index.php'
response = requests.get(url, verify=False)
soup = BeautifulSoup(response.content, 'html.parser')

# Rechercher toutes les tables dans la page
tables = soup.find_all('table')

string_result = ""

# Parcourir chaque table et chercher une correspondance avec des mots-clés
for i, table in enumerate(tables):
    if "PV Grade" in table.get_text() or "Silicon" in table.get_text() or "PV Module" in table.get_text():
        # Parcourir les lignes de cette table
        for row in table.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 7:  # Si il y a assez de colonnes dans la ligne
            
                item = cells[0].get_text(strip=True)
                high_price = cells[1].get_text(strip=True)
                low_price = cells[2].get_text(strip=True)
                average_price = cells[3].get_text(strip=True)
                change = cells[4].get_text(strip=True)
                change_percent = cells[5].get_text(strip=True)
                price_cny = cells[6].get_text(strip=True)
                # Afficher les valeurs extraites
                table_result = {
                    'Item': item,
                    'High Price (USD/kg)': high_price,
                    'Low Price (USD/kg)': low_price,
                    'Average Price (USD/kg)': average_price,
                    'Change (USD)': change,
                    'Change (%)': change_percent,
                    'Price (CNY)': price_cny
                }
                string_result = string_result + str(table_result)
                
usd_cny_rate = re.search(r'USD/CNY\s*:\s*(\d+\.\d+)', string_result).group(1)

# Using regular expression to capture item and prices from the text
pattern = r"'Item':\s*'([^']*)',\s*'High Price \(USD/kg\)':\s*'([^']*)',\s*'Low Price \(USD/kg\)':\s*'([^']*)',\s*'Average Price \(USD/kg\)':\s*'([^']*)',\s*'Change \(USD\)':\s*'([^']*)',\s*'Change \(%\)':\s*'([^']*)',\s*'Price \(CNY\)':\s*'([^']*)'"

matches = re.findall(pattern, string_result)

# Creating a DataFrame from the extracted data
columns = ['Item', 'High Price (USD/kg)', 'Low Price (USD/kg)', 'Average Price (USD/kg)', 'Change (USD)', 'Change (%)', 'Price (CNY)']
df = pd.DataFrame(matches, columns=columns)

# Converting columns to numeric where appropriate
df['High Price (USD/kg)'] = pd.to_numeric(df['High Price (USD/kg)'], errors='coerce')
df['Low Price (USD/kg)'] = pd.to_numeric(df['Low Price (USD/kg)'], errors='coerce')
df['Average Price (USD/kg)'] = pd.to_numeric(df['Average Price (USD/kg)'], errors='coerce')
df['Change (USD)'] = pd.to_numeric(df['Change (USD)'], errors='coerce')
df['Price (CNY)'] = pd.to_numeric(df['Price (CNY)'], errors='coerce')

import openpyxl
from datetime import datetime

# Obtenir la date actuelle sous forme de chaîne au format "YYYY-MM-DD"
current_date = datetime.now().strftime("%Y-%m-%d")
# Nom du fichier avec la date
filename = f'files\pvinsights_{current_date}.xlsx'
if os.path.isfile(filename):
    print("fichier existant")
else:
# Saving the DataFrame to an Excel file
    df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)  # Load your Excel file
    ws = wb.active
    ws["H1"] = "USD/CNY"
    ws["H7"] = usd_cny_rate
    ws.delete_rows(2, 4)
      # Deletes rows 2 to 4 (starting at row 2, delete 3 rows)
    rows_to_delete = [28, 23, 22, 20, 19, 11, 10, 5,4,2]
    
    for row in rows_to_delete:
        ws.delete_rows(row)
    wb.save(filename)  # Save the updated file
    print("SAVED 2")